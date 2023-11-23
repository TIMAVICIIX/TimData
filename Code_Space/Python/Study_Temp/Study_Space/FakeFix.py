import random

import openpyxl

# 打开Excel文件
workbook = openpyxl.load_workbook('C://Users/小吴同志的R7000P/Desktop/WorkSpace.xlsx')

# 选择要操作的工作表
sheet = workbook['Sheet1']  # 将'Sheet1'替换为你的工作表名称

for row in sheet.iter_rows(min_row=2):
    if random.random() < 0.10:  # 10%的几率
        # 在N列插入随机数数据范围在544到568之间
        sheet.cell(row=row[0].row, column=14, value=random.randint(544, 568))

        # 在O列插入随机数范围在9000到11990之间
        sheet.cell(row=row[0].row, column=15, value=random.randint(9000, 11990))

# 保存修改后的Excel文件
workbook.save('C://Users/小吴同志的R7000P/Desktop/WorkSpace.xlsx')
