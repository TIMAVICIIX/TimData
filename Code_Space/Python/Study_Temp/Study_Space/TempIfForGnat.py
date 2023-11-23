import time

import openpyxl
from fuzzywuzzy import fuzz
from openpyxl.styles import Alignment
from selenium import webdriver
from selenium.common import NoSuchElementException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By

# 浏览器驱动显式等待时间
wait_second1 = 3
wait_second2 = 5

# 设置最远爬取年份
end_year = 2021

url = 'https://www.gaokao.cn/lineschool'
# response = requests.get(url)

chrome_options = Options()
# 设置浏览器参数
# --headless是不 显示浏览器启动和执行过程c
chrome_options.add_argument('--headless')
# 设置lang和User-Agent信息，防止反爬虫检测
# 设置本地已存在Chrome浏览器进行自动测试,命令提示符中输入: chrome --remote-debugging-port=9222 并保持浏览器,打开所要爬取的网站
chrome_options.add_argument('lang=zh_CN.UTF-8')
chrome_options.debugger_address = '127.0.0.1:9222'
UserAgent = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/104.0.5112.101 ' \
            'Safari/537.36'
chrome_options.add_argument('User-Agent=' + UserAgent)

# 启动浏览器并设置chrome_options参数
driver = webdriver.Chrome(options=chrome_options)
driver.implicitly_wait(wait_second2)
driver.get(url)

# 数据处理区↓↓↓
'''
1.首先集中数据,缓存大学名称以及代号
2.在TAG中检查二本录取情况,获取招生类型
3.在专业分数线中获取多个年份的录取情况
4.在招生计划中获取各年份的录取情况
整合数据,以大学名称以及代号为主键,索引该学校的各年份信息
以年份为分类,每个年份包含[招生类型],[专业名称],[说明],[最低分/最低位次],[录取率],[计划招生],[学制],[学费]
对招生计划表进行合并时,以专业分数线的专业名称为主,首先遍历专业分数线中有无相同专业名称,再招生计划信息合并
若没有则进行contain包含查找,招生计划中有专业名称是招生计划专业的一部分时,招生计划信息进行合并
'''
# 建立存储数据结构模板
list_info_template = {
    '专业代码': '',
    '专业名称': '',
    '说明': '',
    '再选科目': ' ',
    '23计划': '',
    '22计划': '',
    '21计划': '',
    '学制': '',
    '学费': '',
    '23分数': '',
    '23位次': '',
    '22分数': '',
    '22位次': '',
    '21分数': '',
    '21位次': ''
}

# 存储数据结构如下:
# 建立大学名称以及代号字段
school_name = ''
school_mark = ''
# 建立办学主体字符串
school_type = ''
# 建立主数据结构,访问例子:list_info_all[0]['最低分']
list_info_all = []


# 部署修改大学代号及名称方法
def insert_school_info(name, mark):
    global school_name
    global school_mark

    school_mark = mark
    school_name = name

    print(f'当前学校信息:\n名字:{school_name}\n代号:{school_mark}')


# 部署清除大学代号及名称的方法
def clear_school_info():
    global school_name
    global school_mark

    school_name = ''
    school_mark = ''


# 插入办学主体
def insert_school_type(type_str: str):
    global school_type

    # print(type_str.split(' '))

    temp_school_type = type_str.split(' ')[2]

    if temp_school_type not in school_type:
        school_type = school_type + '[' + temp_school_type + ']\n'


def display_type():
    print(f'办学主体:{school_type}')


# 清除办学主体
def clear_school_type():
    global school_type
    school_type = ''


# 找出专业名称中的说明
def find_tips(tip_str: str):
    track_back = tip_str.find('（')

    if track_back != -1:
        return tip_str[track_back:]
    else:
        print('无说明提取')
        return ''


# 专业名称模糊匹配
def fuzz_patten(patten_str: str):
    # 主动干预匹配精度
    if '（语种：不限）' in patten_str:
        patten_str.replace('（语种：不限）', '')

    patten_str.replace('（', '')
    patten_str.replace('）', '')

    # 创建字典,保存专业名称与匹配度
    name_matches = {}

    # 在每个专业中进行专业名称的模糊匹配
    for match_list in list_info_all:
        # 获取专业名称匹配度
        match_ratio = fuzz.ratio(patten_str, match_list['专业名称'])
        # 填入字典['专业名称':'匹配度']
        name_matches[match_list['专业名称']] = str(match_ratio)

    # Debug
    # print(f'匹配专业:{patten_str}')
    # print(name_matches)

    if len(name_matches) != 0:
        # 返回专业名称匹配度最大的专业名称
        return str(max(name_matches, key=name_matches.get))


# 部署整合大学专业分数线,招生计划的方法
def kit_for_info(plan, strs: str, kit_year):
    # 加入全局变量数据结构
    global list_info_all

    # Debug 专业入参
    # print(f'plan:{plan}\nstrs:{strs}\nkit_year:{kit_year}')

    # 将入参的字符串分割
    into_strs = strs.split('\n')

    # 如果是专业分数线入参，那么应该有5个元素
    if plan == 'part1':
        # 为专业分数线创建新专业子数据结构
        new_temp_list = list_info_template.copy()

        # 为最积分/最低位次进行拆分
        low_score_rank = into_strs[3].split('/')

        # 对列表进行查找如果没有之前创建的专业数据结构则直接进行入参
        check_position = -1
        for check_part_list in list_info_all:
            if check_part_list['专业名称'] == into_strs[0]:
                check_position = list_info_all.index(check_part_list)
                break

        if check_position == -1:
            new_temp_list['专业代码'] = str(len(list_info_all) + 1)
            new_temp_list['专业名称'] = into_strs[0]

            # 匹配模式进行括号内容查找
            track_index = into_strs[0].find('（')

            new_temp_list['说明'] = find_tips(into_strs[0])

            new_temp_list[f'{kit_year}分数'] = low_score_rank[0]
            new_temp_list[f'{kit_year}位次'] = low_score_rank[1]

            # 加入主数据结构
            list_info_all.append(new_temp_list)
            print('新增一个专业!')
        else:
            list_info_all[check_position][f'{kit_year}分数'] = low_score_rank[0]
            list_info_all[check_position][f'{kit_year}位次'] = low_score_rank[1]

        # Debug
        # print(new_temp_list)

    # 如果是招生计划,就从现有数据结构中进行查询
    if plan == 'part2':

        # 声明模糊匹配标记
        find_name = False

        # 遍历查询寻找
        for plan2_struct in list_info_all:
            # 找到相同的专业名称或者具有专业名称包含的数据结构
            if plan2_struct['专业名称'] == into_strs[0]:
                # 不用进行模糊匹配
                find_name = True
                # 进行数据插入
                plan2_struct[f'{kit_year}计划'] = into_strs[1]

                # 匹配模式进行括号内容查找
                plan2_struct['说明'] = find_tips(into_strs[0])

                if plan2_struct['学制'] == '' or plan2_struct['学费'] == '':
                    plan2_struct['学制'] = into_strs[2]
                    plan2_struct['学费'] = into_strs[3]
                break

        # 进行专业名称的模糊匹配
        if not find_name:
            # 模糊匹配获取专业函数名称
            max_ratio_name = fuzz_patten(into_strs[0])

            # 再进行插入
            for part_list in list_info_all:

                if part_list['专业名称'] == max_ratio_name:

                    part_list[f'{kit_year}计划'] = into_strs[1]
                    part_list['说明'] = find_tips(into_strs[0])

                    if part_list['学制'] == '' or part_list['学费'] == '':
                        part_list['学制'] = into_strs[2]
                        part_list['学费'] = into_strs[3]
                    break

            # 被优化,改为模糊匹配
            # elif plan2_struct['专业名称'] in into_strs[0]:
            #     # 进行数据插入
            #     plan2_struct[f'{kit_year}计划'] = into_strs[1]
            #
            #     # 匹配模式进行括号内容查找
            #     plan2_struct['说明'] = find_tips(into_strs[0])
            #
            #     if kit_year == '23':
            #         plan2_struct['学制'] = into_strs[2]
            #         plan2_struct['学费'] = into_strs[3]
            #     break


# 展示大学的所有数据
def display_all_info():
    print('\n信息收集整合完毕')
    print(f'大学名称: {school_name}')
    print(f'大学代号: {school_mark}')
    print(f'大学办学主体: {school_type}')
    print(f'全部专业: ')

    for tmp_print in list_info_all:
        print(tmp_print)


# 部署清除主数据结构
def clear_list_all():
    list_info_all.clear()


# 部署全体数据清除方法
def clear_all_info():
    # 开始清除
    clear_school_info()
    clear_school_type()
    clear_list_all()


# 数据处理区↑↑↑

# 数据保存区↓↓↓

'''
一个大学的数据保存好后,会包含以下信息:
1.大学名称
2.大学代号
3.全部专业的办学主体
4.每个专业的数据结构:
    4.1 专业名称
    4.2 说明
    4.3.1 23计划
    4.3.2 22计划
    4.3.3 21计划
    4.4 学费
    4.5 学制
    4.6.1 23分数
    4.6.2 23位次
    4.7.1 22分数
    4.7.2 22位次
    4.8.1 21分数
    4.8.2 21位次

数据插入顺序:
1.记下最大记录行,统计该学校拥有专业数
2.专业数据结构每个数据结构从第6列进行插入
3.使用最大记录行,定位第一条数据插入位置
4.插入大学主要数据,进行单元格向下合并


'''

# 得到Excel文件实例
excel_inti = openpyxl.load_workbook('C://Users/小吴同志的R7000P/Desktop/WorkSpace.xlsx')
# 设置工作表
excel_sheet = excel_inti["Sheet1"]


# 从第6行进行专业数据结构的插入操作处理
def list_insert():
    # 获取总数据结构
    global list_info_all
    start_row = int(excel_sheet.max_row) + 1
    # 设置起始列
    start_column = 5

    # 主数据结构中循环,提取每一个专业数据结构
    for j, part_list in enumerate(list_info_all):
        # 将专业数据结构的Value值提取出来做一个数组
        part_list_value = list(part_list.values())

        # 该值数组迭代
        for i, value in enumerate(part_list_value):
            # 从开始行迭代插入数据,列从第6列开始迭代
            excel_sheet.cell(row=start_row + int(j), column=start_column + int(i), value=value)


# 回溯进行上下单元格合并并且进行学院数据插入
def main_info_insert():
    # 获取专业数与最大记录行
    max_row = excel_sheet.max_row
    struct_num = len(list_info_all)

    # 设置格式
    # 创建一个 Alignment 对象，设置垂直居中和水平居中
    alignment = Alignment(vertical='center', horizontal='center')

    # 遍历 B、C、D 列的所有单元格，设置垂直居中和水平居中
    for row in excel_sheet.iter_rows(min_row=2, max_row=excel_sheet.max_row, min_col=2, max_col=4):
        for cell in row:
            cell.alignment = alignment

    # 设置列 B、C、D 的宽度
    excel_sheet.column_dimensions['B'].width = 10  # 设置列 B 的宽度为 10
    excel_sheet.column_dimensions['C'].width = 22  # 设置列 C 的宽度为 22
    excel_sheet.column_dimensions['D'].width = 22  # 设置列 D 的宽度为 22
    excel_sheet.column_dimensions['F'].width = 22  # 设置列 F 的宽度为 22
    excel_sheet.column_dimensions['G'].width = 22  # 设置列 G 的宽度为 22

    # 插入主要数据
    excel_sheet.cell(row=max_row - (struct_num - 1), column=2, value=school_mark)
    excel_sheet.cell(row=max_row - (struct_num - 1), column=3, value=school_name)
    excel_sheet.cell(row=max_row - (struct_num - 1), column=4, value=school_type)

    # 开始向下合并单元格
    # excel_sheet.merge_cells(f"A{max_row - (struct_num - 1)}:A{max_row}")
    # excel_sheet.merge_cells(f"B{max_row - (struct_num - 1)}:B{max_row}")
    # excel_sheet.merge_cells(f"C{max_row - (struct_num - 1)}:C{max_row}")
    # excel_sheet.merge_cells(f"D{max_row - (struct_num - 1)}:D{max_row}")


# 保存文件结构,在最后进行
def save_excel():
    excel_inti.save("C://Users/小吴同志的R7000P/Desktop/WorkSpace.xlsx")


# 结构整合,进行总接口建立,上传完毕后自动清空数据结构
def save_bus_and_clear():
    list_insert()
    main_info_insert()
    clear_all_info()
    save_excel()


# 数据保存区↑↑↑

# 提取方法区↓↓↓
# 全局使用abs Xpath

# 二本录取线计数器
TAG_counter = 0
# 年份省级分数线元素定位地址
TAG_year_vis = '/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]/div[1]/div[3]/div[1]' \
               '/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]'
# 年份专业分数线元素定位地址
PLAN_year_vis = '/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]/div[1]/div[3]/div[1]' \
                '/div[1]/div[1]/div[3]/div[1]/div[1]/div[2]'

# 年份招生计划元素定位地址
STU_year_vis = '/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]/div[1]/div[3]/div[1]' \
               '/div[1]/div[1]/div[4]/div[1]/div[1]/div[2]'


# 大学列表提取方法
def link_university():
    time.sleep(1)
    university_lists = driver.find_elements(By.XPATH, '/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/'
                                                      'div[2]/div[2]/div[1]/div[1]/div[1]/div[4]/div[1]/div')
    return university_lists


# 屏幕点击方法
def click_method(path, wait_time):
    time.sleep(wait_time)
    driver.find_element(By.XPATH, path).click()


# 获取text方法
def get_text_method(path, wait_time):
    time.sleep(wait_time)
    return driver.find_element(By.XPATH, path).text


# 获取组件数量方法
def get_com_num(path, wait_time):
    time.sleep(wait_time)

    # Debug
    # for tmp in driver.find_elements(By.XPATH, path):
    #     print(tmp.text)

    return len(driver.find_elements(By.XPATH, path))


# TAG的Table提取方法,不具备翻页能力以及多项选择能力
def get_table_info(path, counter):
    # 时延
    time.sleep(1)

    info_table = driver.find_element(By.XPATH, path)

    tmp_elements = info_table.find_element(By.TAG_NAME, 'tbody').find_elements(By.TAG_NAME, 'tr')

    for tr_element in tmp_elements:

        if tmp_elements.index(tr_element) != 0:
            td_elements = tr_element.find_elements(By.TAG_NAME, 'td')

            combined_str = " ".join([td.text for td in td_elements])

            if "本科二批" in combined_str:
                counter += 1
                # print(combined_str)

                # AOP数据导入,导入三年来的办学主体
                insert_school_type(combined_str)
                # display_type()

            else:
                print('非本科二批信息,不插入\n')

    return counter


# PLAN的提取方法,适用于[专业分数线]与[招生计划]
def get_table_plan(path, plan_current_year, kind):
    # 时延
    time.sleep(1)

    change_page_num = 0
    change_btn_position = 0

    # 专业分数线的提取位置
    if kind == 'part1':
        print('查找任务开始')

        try:
            change_page_num = get_com_num(
                '/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]/div[1]/div[3]/div[1]'
                '/div[1]/div[1]/div[3]/div[2]/div[3]/div[1]/div[1]/ul[1]/li', 0) - 2

            print('列表总页数:' + str(change_page_num))
        except NoSuchElementException:
            print('无列表显示项,列表为1页\n')
            change_page_num = 1

        try:
            change_btn_position = get_com_num(
                '/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]/div[1]/div[3]/div[1]'
                '/div[1]/div[1]/div[3]/div[2]/div[3]/div[1]/div[1]/ul[1]/li', 0)
        except NoSuchElementException:
            print('无下一页按钮,默认为0\n')

    # 招生计划的提取位置
    if kind == 'part2':
        print('查找任务开始')

        try:
            change_page_num = get_com_num(
                '/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]/div[1]'
                '/div[3]/div[1]/div[1]/div[1]/div[4]/div[2]/div[3]/div[1]/div[1]/ul[1]/li', 0) - 2

            print('列表总页数:' + str(change_page_num))
        except NoSuchElementException:
            print('无列表显示项,列表为1页\n')
            change_page_num = 1

        try:
            change_btn_position = get_com_num(
                '/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]/div[1]'
                '/div[3]/div[1]/div[1]/div[1]/div[4]/div[2]/div[3]/div[1]/div[1]/ul[1]/li', 0)
        except NoSuchElementException:
            print('无下一页按钮,默认为0\n')

    print('列表下一页按钮位置:' + str(change_btn_position))

    for i in range(0, change_page_num):
        # 时延
        time.sleep(1)

        info_table = driver.find_element(By.XPATH, path)

        tmp_elements = info_table.find_element(By.TAG_NAME, 'tbody').find_elements(By.TAG_NAME, 'tr')

        for tr_element in tmp_elements:

            if tmp_elements.index(tr_element) != 0:
                td_elements = tr_element.find_elements(By.TAG_NAME, 'td')
                td_info = [td.text for td in td_elements]
                consist_info = '\n'.join(td_info)

                # Debug:get_table_plan()
                # print(consist_info)
                # AOP数据导入,导入专业分数线或招生计划信息
                kit_for_info(kind, consist_info, plan_current_year)

        # 点击Table翻页
        if change_page_num != 1:
            # 专业分数线
            if kind == 'part1' and change_btn_position != 0:
                click_method(
                    '/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]/div[1]/div[3]/div[1]'
                    '/div[1]/div[1]/div[3]/div[2]/div[3]/div[1]/div[1]/ul[1]/li[' +
                    str(change_btn_position) + ']', 0)
            # 招生计划
            if kind == 'part2' and change_btn_position != 0:
                click_method('/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]/div[1]'
                             '/div[3]/div[1]/div[1]/div[1]/div[4]/div[2]/div[3]/div[1]/div[1]/ul[1]/li[' +
                             str(change_btn_position) + ']', 0)


# 更换年份
def change_inner_year(path_open, path_chose):
    time.sleep(1.5)
    driver.find_element(By.XPATH, path_open).click()
    time.sleep(0.5)

    try:
        driver.find_element(By.XPATH, path_chose).click()
    except NoSuchElementException:
        print('无目标年份')


# 退出页面并关闭
def quit_web_page():
    driver.close()
    driver.switch_to.window(driver.window_handles[0])


# 下一页大学列表
def change_main_page():
    list_end = get_com_num('/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]/div[2]'
                           '/div[1]/div[1]/div[1]/div[4]/div[2]/div[1]/ul[1]/li', 2)

    print('底端序号' + str(list_end) + '\n')

    click_method(f'/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]/div[2]/div[1]'
                 f'/div[1]/div[1]/div[4]/div[2]/div[1]/ul[1]/li[{list_end}]', 1)


# 提取方法区↑↑↑

# ！！！代码执行区！！！
# Relate/Tips
# 外循环(对列表展示界面进行遍历)->内循环(对列表展示界面中的各列表项进行遍历)
# 外循环,首先获取列表界面数量
university_list_range = 39

click_method('/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]/div[2]/div[1]'
             '/div[1]/div[1]/div[2]/div[1]/div[1]/div[3]/div[5]/label[2]/span[1]', 1)

click_method('/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]/div[2]'
             '/div[1]/div[1]/div[1]/div[2]/div[1]/div[2]/button[1]', 0.5)

click_method('/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]/div[2]/div[1]'
             '/div[1]/div[1]/div[3]/div[2]/div[3]/div[1]/div[1]', 0.5)

click_method('/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]/div[2]/'
             'div[1]/div[1]/div[1]/div[3]/div[2]/div[3]/div[2]/div[1]/div[1]/div[1]/ul[1]/li[3]', 0.5)

target_page = 7
target_uni = 1

for i in range(1, target_page):
    change_main_page()

for j in range(university_list_range):

    print('列表第' + str(j + 1) + '页\n')
    # 定位单页面界面各大学列表点击按钮
    university_list = link_university()

    # Debug
    print('大学界面列表提取完毕\n')

    for i, university in enumerate(university_list[(target_uni-1):], start=target_uni):
        time.sleep(1)

        un_btn = university.find_element(By.CLASS_NAME, 'line-school_schoolInfo__1sdvn')
        print('\n序号 ' + str(i))
        un_text = un_btn.text
        # driver.execute_script("arguments[0].click();", un_btn)

        # 点进学校之前进行二本院校筛选
        filter_strings = ["985", "211", "专科"]
        if any(substr in un_text for substr in filter_strings):
            print('略过学校:' + un_text)
            continue

        un_btn.click()

        # 当前URL被重新定向到新页面，需要转到新的URL界面
        driver.switch_to.window(driver.window_handles[1])
        print('\n进入大学主界面,当前URL:' + driver.current_url + '\n')

        # AOP数据导出切片,导出大学名称与大学代号并写入数据缓存
        insert_school_info(un_text.split("\n")[0], driver.current_url.split('/')[-1])

        # Debug
        # 对第二部分[专业分数线]的单元测试,测试目标:爬取精度
        # 测试URL: https://www.gaokao.cn/school/[大学代号]
        # 测试结果:测试修改成功,解决Table为一页时任然翻页问题
        # 测试代码如下:
        # driver.execute_script('window.location.href="https://www.gaokao.cn/school/324";')

        # 进入provinceline
        driver.execute_script(f'window.location.href="{driver.current_url}/provinceline";')
        time.sleep(2)
        # click_method('/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[1]'
        #              '/div[1]/div[1]/div[2]/div[1]/div[2]/div[1]/div[5]/div[2]', 4)

        # 滑动屏幕又利于获取表格
        driver.execute_script(f"window.scrollBy(0,400);")

        # 设置TAG中的二本录取情况
        TAG_start_year = 0

        try:

            TAG_start_year = int(get_text_method('/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]'
                                                 '/div[2]/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]/div[1]'
                                                 '/div[1]/div[2]/form[1]/div[1]/div[1]/div[1]/span[1]/div[1]'
                                                 '/div[1]/div[1]/div[1]', 0))
        except ValueError:
            clear_all_info()
            quit_web_page()
            print('无录取情况!跳回')
            continue

        # 设置专业分数线的二本录取情况
        PLAN_start_year = 0

        try:
            PLAN_start_year = int(get_text_method('/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]'
                                                  '/div[2]/div[1]/div[3]/div[1]/div[1]/div[1]/div[3]/div[1]/div[1]'
                                                  '/div[2]/form[1]/div[1]/div[1]/div[1]/span[1]/div[1]/div[1]'
                                                  '/div[1]/div[1]', 0))
        except ValueError:
            print('无专业分数线查找年份!')

        # 设置招生计划的二本录取情况
        STU_start_year = 0

        try:

            # driver.execute_script(f"window.scrollBy(0,1000);")

            # time.sleep(3)

            STU_start_year = int(get_text_method('/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]'
                                                 '/div[1]/div[2]/div[1]/div[3]/div[1]/div[1]/div[1]/div[4]/div[1]'
                                                 '/div[1]/div[2]/form[1]/div[1]/div[1]/div[1]/span[1]/div[1]'
                                                 '/div[1]/div[1]/div[1]', 2))

            # driver.execute_script(f"window.scrollBy(0,-1000);")

        except ValueError:
            print('无招生情况查找年份!')

        # Debug
        # 各信息表的最近年份
        print(
            f'该大学的各信息表最近年份\nTAG:{TAG_start_year}\n专业分数线:{PLAN_start_year}\n招生计划:{STU_start_year}\n')

        # 第一部分:TAG查找开始
        # 获得现在年份,在现在年份与最终年份之间循环查找,并记录有无二本录取情况
        for i in range(2, (TAG_start_year - end_year) + 3):

            TAG_current_year = get_text_method(
                '/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]'
                '/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]'
                '/form[1]/div[1]/div[1]/div[1]/span[1]/div[1]/div[1]/div[1]/div[1]', 0)

            print(f'TAG中当前查找年份:{TAG_current_year}')

            # 选择文科，如果现有文科则不选择，如果没有呈现则击文理科按钮并选择文史，如果没有文科则直接跳过该年份
            if get_text_method('/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]/div[1]/div[3]'
                               '/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]',
                               0) != '文科':

                click_method('/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]'
                             '/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[3]', 1)

                if get_com_num('/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]'
                               '/div[2]/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]'
                               '/div[3]/div[2]/div[1]/div[1]/div[1]/ul[1]/li', 0) == 2:
                    click_method('/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]'
                                 '/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[3]/div[2]'
                                 '/div[1]/div[1]/div[1]/ul[1]/li[2]', 0.5)

                else:
                    print('当前年份无文科信息,切换年份或退出本页面')
                    change_inner_year(TAG_year_vis,
                                      '/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]/div[1]'
                                      '/div[3]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]/form[1]/div[1]'
                                      '/div[1]/div[1]/span[1]/div[2]/div[1]/div[1]/div[1]/ul[1]/li[' + str(i) + ']')
                    continue

            # 进行二本信息查找
            TAG_counter += get_table_info('/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]'
                                          '/div[2]/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]/div[2]/div[1]/table[1]',
                                          TAG_counter)

            print('当前年份文科信息收集完毕,切换年份,Table或退出本页面')
            change_inner_year(TAG_year_vis,
                              '/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]/div[1]'
                              '/div[3]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]/form[1]/div[1]'
                              '/div[1]/div[1]/span[1]/div[2]/div[1]/div[1]/div[1]/ul[1]/li[' + str(i) + ']')

        if TAG_counter == 0:
            print('该校近几年无二本招生情况，退出该页')
            quit_web_page()
            continue
        else:
            TAG_counter = 0

        # 如果有查找年份
        if PLAN_start_year != 0:
            # 进行专业分数线信息查找
            # 第二部分:对专业分数线进行查找,后期可能进行方法抽象优化
            print('\n开始查找专业分数线')
            for i in range(2, (PLAN_start_year - end_year) + 3):

                PLAN1_current_year = get_text_method(
                    '/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]'
                    '/div[1]/div[3]/div[1]/div[1]/div[1]/div[3]/div[1]/div[1]/div[2]/form[1]/div[1]'
                    '/div[1]/div[1]/span[1]/div[1]/div[1]/div[1]/div[1]', 0)

                print(f'\n专业分数线中当前查找年份:{PLAN1_current_year}')

                # 选择文科，如果现有文科则不选择，如果没有呈现则击文理科按钮并选择文史，如果没有文科则直接跳过该年份
                if get_text_method(
                        '/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]/div[1]/div[3]'
                        '/div[1]/div[1]/div[1]/div[3]/div[1]/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]',
                        0) != '文科':

                    click_method('/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]/div[1]'
                                 '/div[3]/div[1]/div[1]/div[1]/div[3]/div[1]/div[1]/div[3]/div[1]', 1)

                    if get_com_num('/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]'
                                   '/div[2]/div[1]/div[3]/div[1]/div[1]/div[1]/div[3]'
                                   '/div[1]/div[1]/div[3]/div[2]/div[1]/div[1]/div[1]/ul[1]/li', 0) == 2:
                        click_method('/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]'
                                     '/div[2]/div[1]/div[3]/div[1]/div[1]/div[1]/div[3]/div[1]/div[1]'
                                     '/div[3]/div[2]/div[1]/div[1]/div[1]/ul[1]/li[2]', 0.5)

                    else:
                        print('选择文科,当前年份无文科信息,切换年份或退出本页面')
                        change_inner_year(PLAN_year_vis,
                                          '/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]'
                                          '/div[2]/div[1]/div[3]/div[1]/div[1]/div[1]/div[3]/div[1]/div[1]'
                                          '/div[2]/form[1]/div[1]/div[1]/div[1]/span[1]/div[2]/div[1]/div[1]'
                                          '/div[1]/ul[1]/li[' + str(i) + ']')
                        continue

                    if get_text_method(
                            '/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]/div[1]'
                            '/div[3]/div[1]/div[1]/div[1]/div[3]/div[1]/div[1]/div[4]/form[1]/div[1]'
                            '/div[1]/div[1]/span[1]/div[1]/div[1]/div[1]/div[1]',
                            0) != '本科二批':

                        click_method('/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]'
                                     '/div[2]/div[1]/div[3]/div[1]/div[1]/div[1]/div[3]/div[1]'
                                     '/div[1]/div[4]/form[1]/div[1]/div[1]/div[1]', 1)

                        if get_com_num('/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]'
                                       '/div[2]/div[1]/div[3]/div[1]/div[1]/div[1]/div[3]/div[1]/div[1]'
                                       '/div[4]/form[1]/div[1]/div[1]/div[1]/span[1]/div[2]/div[1]/div[1]'
                                       '/div[1]/ul[1]/li',
                                       0) == 2:
                            click_method(
                                '/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]/div[1]'
                                '/div[3]/div[1]/div[1]/div[1]/div[3]/div[1]/div[1]/div[4]/form[1]/div[1]/div[1]'
                                '/div[1]/span[1]/div[2]/div[1]/div[1]/div[1]/ul[1]/li[2]', 1)

                        else:
                            print('选择本科批次,当前年份无文科信息,切换年份或退出本页面')
                            change_inner_year(PLAN_year_vis,
                                              '/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[1]'
                                              '/div[1]/div[1]/div[2]/div[1]/div[3]/div[1]/div[1]/div[1]/div[3]'
                                              '/div[1]/div[1]/div[2]/form[1]/div[1]/div[1]'
                                              '/div[1]/span[1]/div[2]/div[1]/div[1]/div[1]/ul[1]/li[' + str(i) + ']')
                            continue

                # 进行二本信息查找
                # driver.execute_script(f"window.scrollBy(0,200);")
                get_table_plan('/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]'
                               '/div[2]/div[1]/div[3]/div[1]/div[1]/div[1]/div[3]/div[2]/div[1]/table[1]',
                               PLAN1_current_year[2:], 'part1')

                driver.execute_script(f"window.scrollBy(0,-250);")
                print('当前年份文科信息收集完毕,切换年份或退出本页面')
                change_inner_year(PLAN_year_vis,
                                  '/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/'
                                  'div[1]/div[1]/div[1]/div[2]/div[1]/div[3]'
                                  '/div[1]/div[1]/div[1]/div[3]/div[1]/div[1]/div[2]/form[1]/div[1]/div[1]'
                                  '/div[1]/span[1]/div[2]/div[1]/div[1]/div[1]/ul[1]/li[' + str(i) + ']')
        else:
            print('专业分数线无年份,不进行查询')

        # 如果有招生计划年份
        if STU_year_vis != 0:
            # 进行招生计划信息查找
            # 第三部分:对招生计划进行查找,后期可能进行方法抽象优化
            print('\n开始查找招生计划')
            for i in range(2, (STU_start_year - end_year) + 3):

                PLAN2_current_year = get_text_method(
                    '/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]/div[1]/div[3]/div[1]'
                    '/div[1]/div[1]/div[4]/div[1]/div[1]/div[2]/form[1]/div[1]/div[1]/div[1]/span[1]/div[1]'
                    '/div[1]/div[1]/div[1]', 0)

                print(f'\n招生计划中当前查找年份:{PLAN2_current_year}')

                # 选择文科，如果现有文科则不选择，如果没有呈现则击文理科按钮并选择文史，如果没有文科则直接跳过该年份
                if get_text_method(
                        '/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]/div[1]'
                        '/div[3]/div[1]/div[1]/div[1]/div[4]/div[1]/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]',
                        0) != '文科':

                    click_method('/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]'
                                 '/div[1]/div[3]/div[1]/div[1]/div[1]/div[4]/div[1]/div[1]/div[3]', 1)

                    if get_com_num('/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]'
                                   '/div[1]/div[3]/div[1]/div[1]/div[1]/div[4]/div[1]/div[1]/div[3]/div[2]'
                                   '/div[1]/div[1]/div[1]/ul[1]/li', 0) == 2:
                        click_method('/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]'
                                     '/div[1]/div[3]/div[1]/div[1]/div[1]/div[4]/div[1]/div[1]/div[3]/div[2]'
                                     '/div[1]/div[1]/div[1]/ul[1]/li[2]', 0.5)

                    else:
                        print('当前年份无文科信息,切换年份或退出本页面')
                        change_inner_year(STU_year_vis,
                                          '/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]'
                                          '/div[1]/div[3]/div[1]/div[1]/div[1]/div[4]/div[1]/div[1]/div[2]/form[1]'
                                          '/div[1]/div[1]/div[1]/span[1]/div[2]'
                                          '/div[1]/div[1]/div[1]/ul[1]/li[' + str(i) + ']')
                        continue

                    if get_text_method(
                            '/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]/div[1]'
                            '/div[3]/div[1]/div[1]/div[1]/div[4]/div[1]/div[1]/div[4]/form[1]'
                            '/div[1]/div[1]/div[1]/span[1]/div[1]/div[1]/div[1]/div[1]',
                            0) != '本科二批':

                        click_method('/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]'
                                     '/div[1]/div[3]/div[1]/div[1]/div[1]/div[4]/div[1]/div[1]/div[4]', 1)

                        if get_com_num('/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]'
                                       '/div[1]/div[3]/div[1]/div[1]/div[1]/div[4]/div[1]/div[1]/div[4]'
                                       '/form[1]/div[1]/div[1]/div[1]/span[1]/div[2]/div[1]/div[1]/div[1]/ul[1]/li',
                                       0) == 2:
                            click_method(
                                '/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]/div[1]'
                                '/div[3]/div[1]/div[1]/div[1]/div[4]/div[1]/div[1]/div[4]/form[1]'
                                '/div[1]/div[1]/div[1]/span[1]/div[2]/div[1]/div[1]/div[1]/ul[1]/li[2]', 1)

                        else:
                            print('当前年份无文科信息,切换年份或退出本页面')
                            change_inner_year(STU_year_vis,
                                              '/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]'
                                              '/div[1]/div[3]/div[1]/div[1]/div[1]/div[4]/div[1]/div[1]/div[2]/form[1]'
                                              '/div[1]/div[1]/div[1]/span[1]/div[2]'
                                              '/div[1]/div[1]/div[1]/ul[1]/li[' + str(i) + ']')
                            continue

                # 进行二本信息查找
                # driver.execute_script(f"window.scrollBy(0,70);")
                get_table_plan('/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]/div[1]'
                               '/div[3]/div[1]/div[1]/div[1]/div[4]/div[2]/div[1]/table[1]', PLAN2_current_year[2:],
                               'part2')

                print('当前年份文科信息收集完毕,切换年份或退出本页面')
                driver.execute_script(f"window.scrollBy(0,-250);")
                change_inner_year(STU_year_vis,
                                  '/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]'
                                  '/div[1]/div[3]/div[1]/div[1]/div[1]/div[4]/div[1]/div[1]/div[2]/form[1]'
                                  '/div[1]/div[1]/div[1]/span[1]/div[2]'
                                  '/div[1]/div[1]/div[1]/ul[1]/li[' + str(i) + ']')

        else:
            print('招生计划无年份,不进行查询')

        display_all_info()
        save_bus_and_clear()
        # Debug
        # clear_all_info()
        # Debug
        # save_excel()
        quit_web_page()

    # 加入翻页功能进入外循环
    change_main_page()

save_excel()
driver.quit()
